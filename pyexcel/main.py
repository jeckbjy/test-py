#!/usr/bin/python3
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import os


def loadUom(file):
    wb = load_workbook(file)
    sh1 = wb['Sheet1']
    r1 = {}
    for index, row in enumerate(sh1.rows):
        if index == 0:
            continue
        b = row[1].value.strip()
        d = int(row[3].value)
        if b not in r1.keys():
            r1[b] = d
        else:
            print("duplicate uom", b, d, r1[b])

    sh2 = wb['Sheet2']
    r2 = {}
    for index, row in enumerate(sh2.rows):
        if index == 0:
            continue
        k = str(row[0].value).strip()
        v = str(row[1].value).strip()
        if k not in r2.keys():
            r2[k] = v
        # else:
        #     print("duplicate HKA", k, v)

    return r1, r2


def toIndex(ch):
    return ord(ch) - ord('A')


def convItemDesc(desc, map):
    (s1, s2, s3) = desc.partition(" ")
    if s1 not in map.keys():
        return s1
    else:
        return map[s1]


def loadData(file, sheet, uom):
    (uom1, uom2) = uom
    result = {}
    wb = load_workbook(file, sheet)
    sh = wb[sheet]
    for index, row in enumerate(sh.rows):
        if index == 0:
            continue
        j = row[toIndex('J')].value
        k = row[toIndex('K')].value
        q = row[toIndex('Q')].value
        t = int(row[toIndex('T')].value)
        u = int(row[toIndex('U')].value)
        v = float(row[toIndex('V')].value)

        if q == 'SG':
            continue

        if q == 'CA' or q == 'CB':
            t = 0
            u = 0

        k = convItemDesc(k, uom2)
        x = uom1[j]
        t *= x
        u *= x

        if k in result.keys():
            old = result[k]
            old[1] += t
            old[2] += u
            old[4] += v
        else:
            result[k] = [k, t, u, 0, v]

    return result


def saveData(file, data):
    wb = Workbook()
    ws = wb.active
    ws.append(['Item Code', 'QtyFoc Pc', 'NetSalesQty Pc','Price 1000Pc', 'NetAmt'])
    for v in data.values():
        if v[2] != 0:
            v[3] = round(v[4] / v[2] * 1000, 2)
        ws.append(v)
    wb.save(file)


if __name__ == '__main__':
    try:
        # input = "15_06_20_sales.xlsx"
        filename = sys.argv[1]
        input = os.path.join("input", filename)
        uom = loadUom("HKA_UOM.XLSX")
        data = loadData(input, "Data", uom)
        outfile = os.path.join("output", filename)
        saveData(outfile, data)
    except Exception as e:
        s = sys.exc_info()
        print("Error '%s' happened on line %d, %s" % (s[1], s[2].tb_lineno, e))
